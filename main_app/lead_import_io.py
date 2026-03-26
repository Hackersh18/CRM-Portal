"""
CSV / Excel row iteration for lead import without pandas (lighter Railway deploys).
"""
from __future__ import annotations

import csv
import io
import math
from typing import Any, Dict, Iterator, Tuple


def is_blank_import_value(val: Any) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    if isinstance(val, str) and not val.strip():
        return True
    return False


def _normalize_header(name: Any) -> str:
    if name is None:
        return ""
    return str(name).strip()


def _iter_csv_rows(file) -> Iterator[Tuple[int, Dict[str, Any]]]:
    raw = file.read()
    if isinstance(raw, bytes):
        text = raw.decode("utf-8-sig")
    else:
        text = str(raw)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return
    # Strip header keys so " email " matches "email"
    rename = {fn: _normalize_header(fn) for fn in reader.fieldnames if fn is not None}
    for i, row in enumerate(reader, start=2):
        out: Dict[str, Any] = {}
        for k, v in row.items():
            if k is None:
                continue
            nk = rename.get(k, _normalize_header(k))
            if not nk:
                continue
            if isinstance(v, str):
                out[nk] = v.strip()
            else:
                out[nk] = v
        yield i, out


def _iter_xlsx_rows(file) -> Iterator[Tuple[int, Dict[str, Any]]]:
    from openpyxl import load_workbook

    data = file.read()
    bio = io.BytesIO(data)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return
        headers = [_normalize_header(h) for h in header_row]
        for i, data_row in enumerate(rows, start=2):
            d: Dict[str, Any] = {}
            for j, val in enumerate(data_row):
                if j < len(headers) and headers[j]:
                    d[headers[j]] = val
            if not any(
                not is_blank_import_value(v) for v in d.values()
            ):
                continue
            yield i, d
    finally:
        wb.close()


def iter_lead_import_rows(file, filename: str) -> Iterator[Tuple[int, Dict[str, Any]]]:
    """
    Yield (row_number, row_dict) for each data row.
    Row numbers start at 2 (row 1 is headers).
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        yield from _iter_csv_rows(file)
    elif name.endswith(".xlsx"):
        yield from _iter_xlsx_rows(file)
    else:
        raise ValueError("Only .csv and .xlsx lead imports are supported.")
